import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
import sys
from extract_transactions_simple import extract_transactions_simple

class PDFToExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF to Excel Converter")
        self.root.geometry("600x400")
        self.root.resizable(True, True)
        
        # Set up the main frame
        main_frame = ttk.Frame(root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Application title
        title_label = ttk.Label(main_frame, text="PDF Table Extractor", font=("Arial", 18, "bold"))
        title_label.pack(pady=10)
        
        # Description
        description = "This tool extracts transaction data from PDF files and saves it to Excel."
        desc_label = ttk.Label(main_frame, text=description, wraplength=500)
        desc_label.pack(pady=5)
        
        # Frame for file selection
        file_frame = ttk.LabelFrame(main_frame, text="Select PDF File", padding="10")
        file_frame.pack(fill=tk.X, pady=10)
        
        # PDF file selection
        self.pdf_path = tk.StringVar()
        pdf_entry = ttk.Entry(file_frame, textvariable=self.pdf_path, width=50)
        pdf_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        browse_btn = ttk.Button(file_frame, text="Browse", command=self.browse_pdf)
        browse_btn.pack(side=tk.RIGHT)
        
        # Frame for output selection
        output_frame = ttk.LabelFrame(main_frame, text="Output Excel File (Optional)", padding="10")
        output_frame.pack(fill=tk.X, pady=10)
        
        # Output file selection
        self.output_path = tk.StringVar()
        output_entry = ttk.Entry(output_frame, textvariable=self.output_path, width=50)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        output_btn = ttk.Button(output_frame, text="Browse", command=self.browse_output)
        output_btn.pack(side=tk.RIGHT)
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, orient=tk.HORIZONTAL, length=100, mode='indeterminate')
        self.progress.pack(fill=tk.X, pady=10)
        
        # Status message
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        status_label = ttk.Label(main_frame, textvariable=self.status_var, wraplength=500)
        status_label.pack(pady=5)
        
        # Convert button
        convert_btn = ttk.Button(main_frame, text="Convert to Excel", command=self.start_conversion)
        convert_btn.pack(pady=10)
        
        # Credits
        credits = "Created for data extraction from financial transaction PDFs"
        credits_label = ttk.Label(main_frame, text=credits, font=("Arial", 8))
        credits_label.pack(side=tk.BOTTOM, pady=5)
    
    def browse_pdf(self):
        filetypes = [("PDF files", "*.pdf")]
        filename = filedialog.askopenfilename(title="Select PDF File", filetypes=filetypes)
        if filename:
            self.pdf_path.set(filename)
            
            # Automatically set a default output path
            base_name = os.path.splitext(os.path.basename(filename))[0]
            output_path = os.path.join(os.path.dirname(filename), f"{base_name}_extraction.xlsx")
            self.output_path.set(output_path)
    
    def browse_output(self):
        filetypes = [("Excel files", "*.xlsx")]
        filename = filedialog.asksaveasfilename(
            title="Save Excel File As",
            filetypes=filetypes,
            defaultextension=".xlsx"
        )
        if filename:
            self.output_path.set(filename)
    
    def start_conversion(self):
        pdf_path = self.pdf_path.get()
        output_path = self.output_path.get() if self.output_path.get() else None
        
        if not pdf_path:
            messagebox.showerror("Error", "Please select a PDF file first!")
            return
        
        if not os.path.exists(pdf_path):
            messagebox.showerror("Error", f"PDF file not found: {pdf_path}")
            return
        
        # Start progress bar
        self.progress.start()
        self.status_var.set("Converting... Please wait.")
        self.root.update()
        
        # Run conversion in a separate thread to keep GUI responsive
        thread = threading.Thread(target=self.run_conversion, args=(pdf_path, output_path))
        thread.daemon = True
        thread.start()
    
    def open_excel_file(self, filepath):
        """Open the Excel file with the default application"""
        import platform
        import subprocess
        
        try:
            if platform.system() == 'Darwin':  # macOS
                subprocess.call(('open', filepath))
            elif platform.system() == 'Windows':
                os.startfile(filepath)
            else:  # Linux
                subprocess.call(('xdg-open', filepath))
        except Exception as e:
            messagebox.showerror("Error", f"Could not open the file: {e}")

    def fix_portfolio_formulas(self, output_path):
        """Fix and add formulas to the portfolio summary section"""
        try:
            import openpyxl
            
            # Load the workbook
            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook.active
            
            # Find Portfolio_Value and Portfolio Summary rows
            portfolio_value_row = None
            portfolio_summary_row = None
            
            for row in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row, column=1).value
                if value == "Portfolio_Value":
                    portfolio_value_row = row
                elif value == "PORTFOLIO SUMMARY":
                    portfolio_summary_row = row
                    break
            
            # Process the Portfolio Summary section
            if portfolio_summary_row:
                # Find header row (should be portfolio_summary_row + 1)
                header_row = portfolio_summary_row + 1
                
                # Verify the headers exist (updated for new column structure)
                if (sheet.cell(row=header_row, column=1).value == "BOM_ID" and
                    sheet.cell(row=header_row, column=2).value == "Scrip_Symbol"):
                    
                    # Process each security row
                    last_row = header_row + 1
                    while last_row <= sheet.max_row and sheet.cell(row=last_row, column=2).value not in (None, "TOTAL"):
                        # The enhanced GOOGLEFINANCE formula is already set during extraction
                        # Just ensure Value formula is correct
                        qty_ref = sheet.cell(row=last_row, column=3).coordinate
                        price_ref = sheet.cell(row=last_row, column=4).coordinate
                        sheet.cell(row=last_row, column=5).value = f"={qty_ref}*{price_ref}"
                        
                        last_row += 1
                    
                    # Process TOTAL row if it exists
                    total_row = last_row
                    if total_row <= sheet.max_row and sheet.cell(row=total_row, column=2).value == "TOTAL":
                        # Update SUM formulas for new column positions
                        first_qty_cell = sheet.cell(row=header_row+1, column=3).coordinate
                        last_qty_cell = sheet.cell(row=total_row-1, column=3).coordinate
                        sheet.cell(row=total_row, column=3).value = f"=SUM({first_qty_cell}:{last_qty_cell})"
                        
                        first_value_cell = sheet.cell(row=header_row+1, column=5).coordinate
                        last_value_cell = sheet.cell(row=total_row-1, column=5).coordinate
                        sheet.cell(row=total_row, column=5).value = f"=SUM({first_value_cell}:{last_value_cell})"
                        
                        # Link Portfolio_Value to the total value (column 5 now)
                        if portfolio_value_row:
                            n_amt_col = None
                            for col in range(1, sheet.max_column + 1):
                                if sheet.cell(row=1, column=col).value == "N.Amt":
                                    n_amt_col = col
                                    break
                            
                            if n_amt_col:
                                total_value_ref = sheet.cell(row=total_row, column=5).coordinate
                                sheet.cell(row=portfolio_value_row, column=n_amt_col).value = f"={total_value_ref}"

                        # Add Portfolio XIRR row after a blank row
                        xirr_row = total_row + 2  # +2 for one blank row
                        sheet.cell(row=xirr_row, column=1).value = "Portfolio XIRR"
                        
                        # Add XIRR formula that uses transaction dates and amounts
                        # Find the date column (3) and N.Amt column
                        n_amt_col = None
                        for col in range(1, sheet.max_column + 1):
                            if sheet.cell(row=1, column=col).value == "N.Amt":
                                n_amt_col = col
                                break
                        
                        if n_amt_col:
                            # Create XIRR formula referencing dates (column 4) and N.Amt
                            # Transaction rows start from row 2 (after header) and go through the Portfolio_Value row
                            date_range = f"D2:D{portfolio_value_row}"
                            amount_range = f"{chr(64+n_amt_col)}2:{chr(64+n_amt_col)}{portfolio_value_row}"
                            # Fix: correct parameter order - values first, then dates
                            xirr_formula = f"=XIRR({amount_range},{date_range})"
                            sheet.cell(row=xirr_row, column=2, value=xirr_formula)
                        
                        # Add Portfolio XIRR Percentage row
                        xirr_pct_row = xirr_row + 1
                        sheet.cell(row=xirr_pct_row, column=1).value = "Portfolio XIRR Percentage"
                        sheet.cell(row=xirr_pct_row, column=2).value = f"={sheet.cell(row=xirr_row, column=2).coordinate}*100"
            
            # Save the workbook
            workbook.save(output_path)
            print(f"Fixed portfolio formulas in {output_path}")
        except Exception as e:
            print(f"Error fixing portfolio formulas: {e}")
    
    def run_conversion(self, pdf_path, output_path):
        try:
            # Redirect stdout to capture console output
            original_stdout = sys.stdout
            from io import StringIO
            captured_output = StringIO()
            sys.stdout = captured_output
            
            # Run the extraction
            result = extract_transactions_simple(pdf_path, output_path)
            
            # Get back console output
            sys.stdout = original_stdout
            log_output = captured_output.getvalue()
            
            # Fix missing Scrip_Symbol values if extraction was successful
            if result is not None:
                actual_output_path = output_path
                if actual_output_path is None:
                    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
                    actual_output_path = f"{base_name}_extraction.xlsx"
                
                # Skip fix_missing_scrip_symbols since it's now handled in extract_transactions_simple.py
                # Just apply portfolio formulas to ensure GOOGLEFINANCE and other calculations work
                self.fix_portfolio_formulas(actual_output_path)
            
            # Update GUI with results
            self.root.after(0, self.update_status, result, output_path, log_output)
            
        except Exception as e:
            self.root.after(0, self.handle_error, str(e))
    
    def update_status(self, result, output_path, log_output):
        self.progress.stop()
        
        if result is not None:
            row_count = len(result)
            if output_path is None:
                base_name = os.path.splitext(os.path.basename(self.pdf_path.get()))[0]
                output_path = f"{base_name}_extraction.xlsx"
                
            success_message = f"Successfully extracted {row_count} rows to {output_path}"
            self.status_var.set(success_message)
            
            messagebox.showinfo("Success", success_message)
            
            # Ask if user wants to open the Excel file
            if messagebox.askyesno("Open Excel File", "Would you like to open the Excel file now?"):
                self.open_excel_file(output_path)
        else:
            error_msg = "Extraction failed. Check if the PDF contains transaction tables."
            self.status_var.set(error_msg)
            messagebox.showerror("Extraction Failed", f"{error_msg}\n\nDetails:\n{log_output}")
    
    def handle_error(self, error_message):
        self.progress.stop()
        self.status_var.set(f"Error: {error_message}")
        messagebox.showerror("Error", f"An error occurred during conversion:\n\n{error_message}")

if __name__ == "__main__":
    root = tk.Tk()
    app = PDFToExcelApp(root)
    root.mainloop()